from __future__ import annotations
from io import BytesIO
import json
from datetime import datetime, date, time
from pathlib import Path
import math

import pandas as pd
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DEV = 'Ing. Christian Pocol, Ingeniero Electrónico'


def _is_missing(value) -> bool:
    """Evalúa nulos sin romper con listas/diccionarios/arrays."""
    if value is None:
        return True
    try:
        if value is pd.NaT:
            return True
    except Exception:
        pass
    if isinstance(value, float) and math.isnan(value):
        return True
    return False


def _safe_cell(value):
    """Convierte cualquier valor de Canvas/Pandas a algo que Excel acepte.

    Esta función evita el error típico:
    "Excel does not support datetimes with timezones".
    También evita problemas cuando Canvas devuelve listas/diccionarios.
    """
    if _is_missing(value):
        return ''

    # Pandas Timestamp, incluso con zona horaria
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return ''
        # Lo mandamos como texto local sin zona para que Excel no lo rechace.
        try:
            if value.tzinfo is not None:
                value = value.tz_convert(None)
        except Exception:
            try:
                value = value.tz_localize(None)
            except Exception:
                pass
        return value.strftime('%Y-%m-%d %H:%M:%S')

    # datetime de Python con tzinfo
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.replace(tzinfo=None)
        return value.strftime('%Y-%m-%d %H:%M:%S')

    # date/time de Python
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, time):
        return value.replace(tzinfo=None).strftime('%H:%M:%S') if value.tzinfo else value.strftime('%H:%M:%S')

    # Estructuras devueltas por Canvas
    if isinstance(value, (dict, list, tuple, set)):
        try:
            return json.dumps(value, ensure_ascii=False, default=str)
        except Exception:
            return str(value)

    # Objetos numpy/pandas raros
    try:
        if hasattr(value, 'item'):
            value = value.item()
    except Exception:
        pass

    # Cadenas ISO de Canvas: se dejan como texto seguro y legible
    if isinstance(value, str):
        return value.replace('T', ' ').replace('Z', '')

    return value


def _write_df(ws, df: pd.DataFrame | None):
    """Escritura manual con openpyxl para no depender de pandas.to_excel."""
    if df is None or df.empty:
        ws.append(['Sin datos disponibles'])
        return

    clean = df.copy()
    clean.columns = [str(c) for c in clean.columns]
    ws.append(list(clean.columns))

    for row in clean.itertuples(index=False, name=None):
        ws.append([_safe_cell(v) for v in row])


def _format_ws(ws):
    ws.freeze_panes = 'A2'
    header_fill = PatternFill('solid', fgColor='003865')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        values = [str(c.value) if c.value is not None else '' for c in column_cells]
        max_len = min(max([len(v) for v in values] + [10]) + 2, 55)
        ws.column_dimensions[column_cells[0].column_letter].width = max_len


def excel_bytes(summary: pd.DataFrame, submissions: pd.DataFrame, history: pd.DataFrame | None = None) -> bytes:
    """Genera el Excel sin usar pandas.to_excel para evitar errores por timezone."""
    bio = BytesIO()
    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'Resumen estudiantes'
    _write_df(ws1, summary)

    ws2 = wb.create_sheet('Entregas detalle')
    _write_df(ws2, submissions)

    if history is not None and not history.empty:
        ws3 = wb.create_sheet('Historial')
        _write_df(ws3, history)

    for ws in wb.worksheets:
        _format_ws(ws)

    wb.save(bio)
    return bio.getvalue()


def _watermark(canvas, doc):
    canvas.saveState()
    canvas.setFont('Helvetica-Bold', 42)
    canvas.setFillColor(colors.Color(0.85, 0.85, 0.85, alpha=0.25))
    canvas.translate(5.5 * inch, 4.2 * inch)
    canvas.rotate(35)
    canvas.drawCentredString(0, 0, 'AVE - UVG')
    canvas.restoreState()
    canvas.saveState()
    canvas.setFont('Helvetica', 8)
    canvas.setFillColor(colors.grey)
    canvas.drawString(0.4 * inch, 0.3 * inch, f'Desarrollador: {DEV}')
    canvas.drawRightString(10.6 * inch, 0.3 * inch, f'Página {doc.page}')
    canvas.restoreState()


def pdf_bytes(summary: pd.DataFrame, course_name: str, section_name: str, generated_by: str, analysis_date: str, logo_ave='assets/logo_ave.png', logo_uvg='assets/logo_uvg.png') -> bytes:
    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Small', fontSize=8, leading=10))
    styles.add(ParagraphStyle(name='TitleAVE', fontSize=18, leading=22, alignment=1, textColor=colors.HexColor('#003865')))
    story = []

    header_data = []
    if Path(logo_ave).exists():
        header_data.append(Image(logo_ave, width=1.2 * inch, height=0.55 * inch))
    else:
        header_data.append('')
    header_data.append(Paragraph('Informe Ejecutivo de Seguimiento Académico AVE', styles['TitleAVE']))
    if Path(logo_uvg).exists():
        header_data.append(Image(logo_uvg, width=1.1 * inch, height=0.55 * inch))
    else:
        header_data.append('')

    ht = Table([header_data], colWidths=[1.4 * inch, 7.2 * inch, 1.4 * inch])
    ht.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
    story += [ht, Spacer(1, 10)]
    story.append(Paragraph(f'<b>Curso:</b> {course_name}<br/><b>Sección:</b> {section_name}<br/><b>Fecha de análisis:</b> {analysis_date}<br/><b>Generado por:</b> {generated_by or "No especificado"}<br/><b>Desarrollador:</b> {DEV}', styles['Small']))
    story.append(Spacer(1, 12))

    total = len(summary)
    counts = summary['riesgo_integral'].value_counts().to_dict() if not summary.empty and 'riesgo_integral' in summary.columns else {}
    kpi = [
        ['Total estudiantes', 'Riesgo bajo', 'Riesgo medio', 'Riesgo alto', 'Prom. avance', 'Prom. horas'],
        [
            total,
            counts.get('Bajo', 0),
            counts.get('Medio', 0),
            counts.get('Alto', 0),
            f"{summary['porcentaje_avance'].mean():.1f}%" if total and 'porcentaje_avance' in summary.columns else '0%',
            f"{summary['tiempo_total_horas'].mean():.1f}" if total and 'tiempo_total_horas' in summary.columns else '0',
        ],
    ]
    kt = Table(kpi, colWidths=[1.5 * inch] * 6)
    kt.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003865')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('GRID', (0, 0), (-1, -1), 0.25, colors.grey), ('ALIGN', (0, 0), (-1, -1), 'CENTER')]))
    story += [kt, Spacer(1, 12)]

    cols = ['estudiante', 'correo', 'horas_sin_actividad', 'riesgo_desconexion', 'tiempo_total_horas', 'pendientes', 'atrasadas', 'porcentaje_avance', 'puntaje_riesgo', 'riesgo_integral', 'accion_recomendada']
    summary_view = summary.copy()
    for col in cols:
        if col not in summary_view.columns:
            summary_view[col] = ''
    view = summary_view[cols].head(35).copy() if not summary_view.empty else pd.DataFrame(columns=cols)
    view.columns = ['Estudiante', 'Correo', 'Horas sin act.', 'Riesgo conexión', 'Horas', 'Pend.', 'Atras.', 'Avance %', 'Puntaje', 'Riesgo total', 'Acción']
    data = [list(view.columns)] + view.fillna('').astype(str).values.tolist()
    table = Table(data, repeatRows=1, colWidths=[1.45 * inch, 1.55 * inch, 0.7 * inch, 0.95 * inch, 0.55 * inch, 0.45 * inch, 0.45 * inch, 0.6 * inch, 0.55 * inch, 0.75 * inch, 1.35 * inch])
    table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6c757d')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey), ('FONTSIZE', (0, 0), (-1, -1), 6.5), ('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    story.append(table)
    story.append(Spacer(1, 8))
    story.append(Paragraph('Nota: el listado prioriza los estudiantes con mayor puntaje de riesgo. Los datos dependen de los permisos del token y de los registros disponibles en Canvas.', styles['Small']))
    doc.build(story, onFirstPage=_watermark, onLaterPages=_watermark)
    return bio.getvalue()
